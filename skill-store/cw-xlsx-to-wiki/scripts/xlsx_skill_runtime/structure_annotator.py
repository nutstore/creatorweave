"""Create an annotated workbook that visualizes inferred structure regions."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from xlsx_skill_runtime.wiki_files import read_yaml_file

ANNOTATED_WORKBOOK_SUFFIX: str = ".structured-annotations.xlsx"
BLOCK_BORDER_COLORS: tuple[str, ...] = (
    "FFC65911",
    "FF2F75B5",
    "FF548235",
    "FFC00000",
    "FF7030A0",
    "FFBF9000",
)


def generate_structural_annotation_workbook(
    *,
    workspace_path: Path,
    source_workbook_path: Path,
    output_path: Path,
) -> tuple[Path, list[str]]:
    """Create a workbook copy with structural regions highlighted."""

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side
    except ModuleNotFoundError as error:  # pragma: no cover - depends on runtime env
        raise RuntimeError(
            "openpyxl is required to generate the structural annotation workbook."
        ) from error

    workbook = load_workbook(filename=source_workbook_path)
    workbook_payload: dict[str, Any] = read_yaml_file(
        path=workspace_path / "wiki" / "workbook.yaml"
    )
    sheet_registry: dict[str, Any] = dict(workbook_payload.get("sheets", {}))
    warnings: list[str] = []

    for block_index, sheet_slug in enumerate(workbook_payload.get("sheet_order", [])):
        sheet_metadata: dict[str, Any] = dict(sheet_registry.get(str(sheet_slug), {}))
        sheet_name: str = str(sheet_metadata.get("name", "")).strip()
        if sheet_name == "" or sheet_name not in workbook.sheetnames:
            warnings.append(
                f"Skipped structural annotation for sheet `{sheet_slug}` because the worksheet was not found in the copied workbook."
            )
            continue
        structure_path: Path = (
            workspace_path / "wiki" / "sheets" / str(sheet_slug) / "structure.yaml"
        )
        if not structure_path.exists():
            warnings.append(
                f"Skipped structural annotation for sheet `{sheet_slug}` because structure.yaml was not found."
            )
            continue
        worksheet = workbook[sheet_name]
        structure_payload: dict[str, Any] = read_yaml_file(path=structure_path)
        regions_index_path: Path = (
            workspace_path / "wiki" / "sheets" / str(sheet_slug) / "regions.yaml"
        )
        region_payloads: list[dict[str, Any]] = _load_region_payloads(
            sheet_dir=workspace_path / "wiki" / "sheets" / str(sheet_slug),
            regions_index_path=regions_index_path,
        )
        warnings.extend(
            _annotate_worksheet(
                worksheet=worksheet,
                structure_payload=structure_payload,
                region_payloads=region_payloads,
                block_palette_offset=block_index,
                border_cls=Border,
                side_cls=Side,
            )
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path, warnings


def _annotate_worksheet(
    *,
    worksheet: Any,
    structure_payload: dict[str, Any],
    region_payloads: list[dict[str, Any]],
    block_palette_offset: int,
    border_cls: Any,
    side_cls: Any,
) -> list[str]:
    warnings: list[str] = []
    block_payloads: list[dict[str, Any]] = list(region_payloads)
    if len(block_payloads) == 0:
        data_region: dict[str, Any] = dict(structure_payload.get("data_region", {}))
        normalized_data_region: dict[str, int] | None = _normalize_bbox(data_region)
        if normalized_data_region is not None:
            block_payloads.append(
                {
                    "block_id": "data_region",
                    "type": "data_region",
                    "source": str(structure_payload.get("data_region_source", "unknown")),
                    "bbox": normalized_data_region,
                }
            )

    for index, block_payload in enumerate(block_payloads):
        bbox: dict[str, int] | None = _normalize_bbox(block_payload.get("bbox", {}))
        if bbox is None:
            continue
        palette_index: int = (block_palette_offset + index) % len(BLOCK_BORDER_COLORS)
        border_color: str = BLOCK_BORDER_COLORS[palette_index]
        _apply_outline_to_bbox(
            worksheet=worksheet,
            bbox=bbox,
            side=side_cls(style="double", color=border_color),
            border_cls=border_cls,
        )

    return warnings


def _load_region_payloads(
    *,
    sheet_dir: Path,
    regions_index_path: Path,
) -> list[dict[str, Any]]:
    if not regions_index_path.exists():
        return []
    regions_index_payload: dict[str, Any] = read_yaml_file(path=regions_index_path)
    loaded_regions: list[dict[str, Any]] = []
    for region_entry in list(regions_index_payload.get("regions", [])):
        embedded_meta: dict[str, Any] = dict(region_entry.get("meta", {}))
        if len(embedded_meta) > 0:
            loaded_regions.append(embedded_meta)
        else:
            loaded_regions.append(dict(region_entry))
    return loaded_regions


def _normalize_bbox(raw_bbox: Any) -> dict[str, int] | None:
    bbox: dict[str, Any] = dict(raw_bbox or {})
    start_row: int = int(bbox.get("start_row", 0))
    end_row: int = int(bbox.get("end_row", 0))
    start_col: int = int(bbox.get("start_col", 0))
    end_col: int = int(bbox.get("end_col", 0))
    if start_row <= 0 or end_row <= 0 or start_col <= 0 or end_col <= 0:
        return None
    if end_row < start_row or end_col < start_col:
        return None
    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
    }


def _apply_outline_to_bbox(
    *,
    worksheet: Any,
    bbox: dict[str, int],
    side: Any,
    border_cls: Any,
) -> None:
    start_row: int = int(bbox["start_row"])
    end_row: int = int(bbox["end_row"])
    start_col: int = int(bbox["start_col"])
    end_col: int = int(bbox["end_col"])
    for row_index in range(start_row, end_row + 1):
        _try_add_border(
            worksheet=worksheet,
            row_index=row_index,
            column_index=start_col,
            border_cls=border_cls,
            left=side,
        )
        _try_add_border(
            worksheet=worksheet,
            row_index=row_index,
            column_index=end_col,
            border_cls=border_cls,
            right=side,
        )
    for column_index in range(start_col, end_col + 1):
        _try_add_border(
            worksheet=worksheet,
            row_index=start_row,
            column_index=column_index,
            border_cls=border_cls,
            top=side,
        )
        _try_add_border(
            worksheet=worksheet,
            row_index=end_row,
            column_index=column_index,
            border_cls=border_cls,
            bottom=side,
        )
def _try_add_border(
    *,
    worksheet: Any,
    row_index: int,
    column_index: int,
    border_cls: Any,
    left: Any | None = None,
    right: Any | None = None,
    top: Any | None = None,
    bottom: Any | None = None,
) -> None:
    try:
        cell = worksheet.cell(row=row_index, column=column_index)
        current_border = cell.border
        cell.border = border_cls(
            left=left or current_border.left,
            right=right or current_border.right,
            top=top or current_border.top,
            bottom=bottom or current_border.bottom,
            diagonal=current_border.diagonal,
            diagonal_direction=current_border.diagonal_direction,
            vertical=current_border.vertical,
            horizontal=current_border.horizontal,
            diagonalUp=current_border.diagonalUp,
            diagonalDown=current_border.diagonalDown,
            outline=current_border.outline,
            start=current_border.start,
            end=current_border.end,
        )
    except AttributeError:
        return